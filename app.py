import streamlit as st
import pandas as pd
import re
import traceback
import io
from openpyxl.styles import PatternFill, Border, Side

st.set_page_config(page_title="仕訳フォーマット自動変換", layout="wide")
st.title("📄 楽楽請求 仕訳フォーマット自動変換ツール")
st.write("会計ソフトから出力した過去仕訳データ(CSV)をアップロードするだけで、楽楽請求の設定フォーマットに変換します！")

uploaded_file = st.file_uploader("1. ここにCSVファイルをドラッグ＆ドロップしてください", type=['csv'])

if uploaded_file is not None:
    st.write(f"2. 「{uploaded_file.name}」 を読み込んでいます...")
    
    df = None
    encodings = ['cp932', 'utf-8', 'utf-8-sig', 'shift_jis']
    for enc in encodings:
        try:
            uploaded_file.seek(0)
            df = pd.read_csv(uploaded_file, encoding=enc)
            st.success(f"文字コード {enc} で読み込み成功！")
            break
        except Exception:
            continue
            
    if df is None:
        st.error("【エラー】CSVファイルの読み込みに失敗しました。")
        st.stop()

    def find_col(keywords):
        for k in keywords:
            for col in df.columns:
                if k == str(col): return col
        for k in keywords:
            for col in df.columns:
                if k in str(col): return col
        return None

    def clean_code(val):
        if pd.isna(val): return ""
        v = str(val).strip()
        if v.endswith('.0'): v = v[:-2]
        if v == 'nan': return ""
        return v

    col_date = find_col(['日付', '伝票日付', '仕訳日'])
    col_denpyo = find_col(['伝票', '伝番', 'No', '番号'])
    col_kari_money = find_col(['本体金額', '借方金額／貸方金額', '借方金額', '借金額', '金額(借)', '金額'])
    col_tekiyo = find_col(['元帳摘要', '摘要', '詳細', 'メモ'])
    
    col_kari_kamoku_name = find_col(['借方勘定科目名', '借方科目名', '借方科目', '借科目', '勘定科目(借)'])
    col_torihikisaki_name = find_col(['取引先名', '貸方補助科目名', '貸方補助', '補助科目名'])

    col_torihikisaki_code = find_col(['取引先コード'])
    col_kari_kamoku_code = find_col(['借方勘定科目コード', '借方科目コード'])
    col_kari_hojo_code = find_col(['借方補助科目コード', '借方補助コード', '借方補助'])
    col_kari_bumon_code = find_col(['借方部門コード', '部門'])
    col_kari_project_code = find_col(['借方プロジェクトコード', 'プロジェクトコード', 'プロジェクト'])
    col_kari_tax = find_col(['借方消費税区分コード', '借方消費税', '借方税区分', '税区分', '税', '消費税'])
    
    col_kashi_kamoku_code = find_col(['貸方勘定科目コード', '貸方科目コード'])
    col_kashi_hojo_code = find_col(['貸方補助科目コード', '貸方補助コード', '貸方補助'])
    col_kashi_bumon_code = find_col(['貸方部門コード'])
    col_kashi_project_code = find_col(['貸方プロジェクトコード'])
    col_kashi_tax = find_col(['貸方消費税区分コード', '貸方消費税', '貸方税区分'])
    
    col_kashi_money = find_col(['本体金額.1', '貸方金額', '貸金額', '金額(貸)'])
    if not col_kashi_money:
        col_kashi_money = col_kari_money 

    if not col_denpyo or not col_kari_kamoku_name:
        st.error(f"【エラー】必須となる「伝票番号」または「借方科目」の列が見つかりません。")
        st.write("▼現在のCSVの列名一覧")
        st.write(list(df.columns))
        st.stop()

    def replace_month(text):
        if pd.isna(text): return ""
        return re.sub(r'[0-9０-９]{1,2}月', '##__請求月__##', str(text))

    try:
        st.write("3. データを変換・整形しています...")
        grouped = df.groupby(col_denpyo)
        patterns = []
        
        for denpyo, group in grouped:
            first_row = group.iloc[0]
            
            t_code = clean_code(first_row[col_torihikisaki_code]) if col_torihikisaki_code else ""
            t_name = str(first_row[col_torihikisaki_name]).replace('nan', '不明').strip() if col_torihikisaki_name else "不明"
            
            is_month_end = True
            if col_date:
                clean_dates = group[col_date].astype(str).str.replace('*', '').str.strip()
                dates = pd.to_datetime(clean_dates, errors='coerce', format='%m.%d')
                if dates.isna().all():
                    dates = pd.to_datetime(clean_dates, errors='coerce')
                
                if dates.isna().all() or not dates.dropna().dt.is_month_end.all():
                    is_month_end = False
            else:
                is_month_end = False

            kari_kamoku_list = group[col_kari_kamoku_name].dropna().unique().tolist()
            if not kari_kamoku_list: kari_kamoku_list = ["科目不明"]
                
            if len(kari_kamoku_list) == 1:
                pattern_name = f"{t_name}_{kari_kamoku_list[0]}"
            elif len(kari_kamoku_list) == 2:
                pattern_name = f"{t_name}_{kari_kamoku_list[0]}/{kari_kamoku_list[1]}"
            else:
                pattern_name = f"{t_name}_{kari_kamoku_list[0]}ほか"

            lines = []
            for index, row in group.iterrows():
                lines.append({
                    'kari_kamoku': clean_code(row[col_kari_kamoku_code]) if col_kari_kamoku_code else "",
                    'kari_hojo': clean_code(row[col_kari_hojo_code]) if col_kari_hojo_code else "",
                    'kari_bumon': clean_code(row[col_kari_bumon_code]) if col_kari_bumon_code else "",
                    'kari_project': clean_code(row[col_kari_project_code]) if col_kari_project_code else "",
                    'kari_tax': clean_code(row[col_kari_tax]) if col_kari_tax else "",
                    'kari_money_val': clean_code(row[col_kari_money]) if col_kari_money else "",
                    'kashi_kamoku': clean_code(row[col_kashi_kamoku_code]) if col_kashi_kamoku_code else "",
                    'kashi_hojo': clean_code(row[col_kashi_hojo_code]) if col_kashi_hojo_code else "",
                    'kashi_bumon': clean_code(row[col_kashi_bumon_code]) if col_kashi_bumon_code else "",
                    'kashi_project': clean_code(row[col_kashi_project_code]) if col_kashi_project_code else "",
                    'kashi_tax': clean_code(row[col_kashi_tax]) if col_kashi_tax else "",
                    'kashi_money_val': clean_code(row[col_kashi_money]) if col_kashi_money else "",
                    'tekiyo': replace_month(row[col_tekiyo]) if col_tekiyo else ""
                })
                
            patterns.append({
                'name': pattern_name,
                't_code': t_code,
                'is_month_end': is_month_end,
                'lines': lines
            })

        unique_patterns_dict = {}
        for p in patterns:
            # ★修正箇所：同じ仕訳かどうかの判定基準（sig）の先頭に、取引先コードを追加しました！
            # これにより、取引先コードが違えば完全に別パターンとして扱われます。
            sig = str(p['t_code']) + "_" + p['name'] + "".join([str(l['kari_kamoku'])+str(l['kashi_kamoku']) for l in p['lines']])
            
            if sig not in unique_patterns_dict:
                unique_patterns_dict[sig] = p
            else:
                unique_patterns_dict[sig]['is_month_end'] = unique_patterns_dict[sig]['is_month_end'] and p['is_month_end']
        
        unique_patterns = list(unique_patterns_dict.values())

        t_code_totals = {}
        for p in unique_patterns:
            t = p['t_code']
            if t: t_code_totals[t] = t_code_totals.get(t, 0) + 1
            
        t_code_current = {}
        output_rows = []
        
        for p in unique_patterns:
            t = p['t_code']
            if not t:
                t_code_current['EMPTY'] = t_code_current.get('EMPTY', 0) + 1
                pattern_code = f"CODE_{t_code_current['EMPTY']}"
            else:
                t_code_current[t] = t_code_current.get(t, 0) + 1
                if t_code_totals[t] > 1:
                    pattern_code = f"{t}_{chr(64 + t_code_current[t])}"
                else:
                    pattern_code = t
            
            date_val = "請求日の当月月末" if p['is_month_end'] else "請求日の当日"
            
            for line in p['lines']:
                row_1 = {
                    '仕訳パターンコード': pattern_code, '仕訳パターン名': p['name'], '取引先': p['t_code'], 'プロジェクト': '', '部門': '', '仕訳日': date_val,
                    '借方_勘定科目_1': '', '借方_補助科目_1': '', '借方_税区分_1': '', '借方_金額_1': '', '借方_部門_1': '', '借方_プロジェクト_1': '',
                    '貸方_勘定科目_1': '', '貸方_補助科目_1': '', '貸方_税区分_1': '', '貸方_金額_1': '', '貸方_部門_1': '', '貸方_プロジェクト_1': '',
                    '共通_摘要_1': ''
                }
                
                row_2 = row_1.copy()
                row_3 = row_1.copy()
                
                for k in row_2.keys():
                    row_2[k] = ""
                    row_3[k] = ""
                    
                row_2['仕訳パターンコード'] = '仕訳の作成方法'
                row_2['仕訳パターン名'] = '請求金額(源泉徴収税額控除前)から作成する'
                
                row_3['仕訳パターンコード'] = '金額の設定方法'
                row_3['仕訳パターン名'] = '金額を直接入力する'
                
                row_1['借方_勘定科目_1'] = line['kari_kamoku']
                row_1['借方_補助科目_1'] = line['kari_hojo']
                row_1['借方_税区分_1'] = line['kari_tax']
                row_1['借方_部門_1'] = line['kari_bumon']
                row_1['借方_プロジェクト_1'] = line['kari_project']
                
                row_1['貸方_勘定科目_1'] = line['kashi_kamoku']
                row_1['貸方_補助科目_1'] = line['kashi_hojo']
                row_1['貸方_税区分_1'] = line['kashi_tax']
                row_1['貸方_部門_1'] = line['kashi_bumon']
                row_1['貸方_プロジェクト_1'] = line['kashi_project']
                
                row_1['共通_摘要_1'] = line['tekiyo']
                
                if len(p['lines']) == 1:
                    row_1['借方_金額_1'] = "差額を自動入力"
                    row_1['貸方_金額_1'] = "差額を自動入力"
                else:
                    row_1['借方_金額_1'] = "固定で金額を入力"
                    row_1['貸方_金額_1'] = "固定で金額を入力"
                    row_2['借方_金額_1'] = line['kari_money_val']
                    row_2['貸方_金額_1'] = line['kashi_money_val']
                    
                output_rows.extend([row_1, row_2, row_3])
            
        final_df = pd.DataFrame(output_rows)
        
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            final_df.to_excel(writer, index=False)
            
            worksheet = writer.sheets['Sheet1']
            color_fill = PatternFill(start_color='E6EDF5', end_color='E6EDF5', fill_type='solid')
            thin_border = Border(
                left=Side(style='thin', color='000000'),
                right=Side(style='thin', color='000000'),
                top=Side(style='thin', color='000000'),
                bottom=Side(style='thin', color='000000')
            )
            
            for col_idx in range(1, len(final_df.columns) + 1):
                worksheet.cell(row=1, column=col_idx).border = thin_border
            
            for row_idx in range(2, len(final_df) + 2):
                block_idx = (row_idx - 2) // 3
                is_colored = (block_idx % 2 != 0)
                
                for col_idx in range(1, len(final_df.columns) + 1):
                    cell = worksheet.cell(row=row_idx, column=col_idx)
                    cell.border = thin_border
                    if is_colored:
                        cell.fill = color_fill

        output.seek(0)
        
        st.success("✅ 変換が完了しました！下のボタンからダウンロードしてください。")
        st.download_button(
            label="📥 フォーマットをダウンロード",
            data=output,
            file_name='rakuraku_format.xlsx',
            mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )

    except Exception as e:
        st.error("【エラー】変換処理中に問題が発生しました。")
        st.text(traceback.format_exc())
